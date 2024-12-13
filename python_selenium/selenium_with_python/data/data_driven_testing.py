from selenium import webdriver
from selenium.webdriver.support.select import Select
import openpyxl
"""
data driven testing - we use this testing when we have some test data and u have to execute multiple time fo same test cases
can prepare data in excel,database
testing is done by test data
selenium - not supported excel
so use third party module - openpyxl - we can work with excel file(.xlsx)
1) how to read data from excel
2) how to write data into excel
3) data case
workbook.active - get active sheet from excel
file -> workbook -> sheets -> rows -> cells
"""
# program to extract data from excel
file = "C:\data\excel.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"]
#find no of rows and column
rows = sheet.max_row
cols = sheet.max_column

for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(r,c).value,end="          ")
    print()

# # write data into excel - for same data in all cells
file = "C:\data\excel2.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"] #or wokbook.active
for r in range(1,6):
    for c in range(1,4):
        sheet.cell(r,c).value="welcome"
workbook.save(file)

# write data into excel - for multiple data in all cells
file = "C:\data\excel3.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"] #or wokbook.active

sheet.cell(1,1).value=123
sheet.cell(1,2).value = "rohini"
sheet.cell(1,3).value = "product manager"

workbook.save(file)

# data driven testing
"""
automation code

excel related operation
read
writing
rows
column
green/red

excelutils :-> we use this utility for every automation test case
read()
writing()
rows()
column()
green/red()
Automation - validation

"""




driver= webdriver.Chrome()